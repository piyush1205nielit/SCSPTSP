"""
Corrected upload view and supporting helpers.
Quarterly filtering now works based on trained_date and certified_date, not session.
"""

import json
import re
import traceback
from datetime import datetime
from decimal import Decimal
from statistics import mode

import openpyxl
from django.contrib import messages
from django.contrib.admin import options
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.views.decorators.csrf import ensure_csrf_cookie
from openpyxl.styles import Font, PatternFill

from .forms import ExcelUploadForm, PlacementUploadForm, StudentDataForm
from .models import Dlc, NsqfElectronics, NsqfIT, PlacementRecord, studentdata, UserProfile

MONTH_MAP = {
    m: i
    for i, m in enumerate(
        [
            "JAN",
            "FEB",
            "MAR",
            "APR",
            "MAY",
            "JUN",
            "JUL",
            "AUG",
            "SEP",
            "OCT",
            "NOV",
            "DEC",
        ],
        1,
    )
}

SORTABLE_FIELDS = {
    "roll_number",
    "batch_code",
    "name",
    "course_name",
    "father_name",
    "mother_name",
    "dob",
    "gender",
    "address",
    "qualifications",
    "aadhaar",
    "scheme",
    "nsqf",
    "course_hour",
    "course_category",
    "center_name",
    "mode",
    "caste_category",
    "fee",
    "claimable_amount",
    "fee_date",
    "trained",
    "certified",
    "placed",
    "session",
    "claimed",
}

CENTERS = ["inderlok", "janakpuri", "karkardooma"]


def get_student_qs(user):
    try:
        profile = user.profile
        if profile.center:
            return studentdata.objects.filter(center_name=profile.center)
    except UserProfile.DoesNotExist:
        pass
    return studentdata.objects.all()


def parse_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in ["true", "yes", "1"]


def parse_date(value):
    """Return a date object or None. Accepts openpyxl date/datetime, or common string formats."""
    if not value:
        return None
    # openpyxl may give datetime/date objects
    if hasattr(value, "date"):
        try:
            return value.date()
        except Exception:
            pass
    val_str = str(value).strip()
    # try multiple formats
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(val_str, fmt).date()
        except Exception:
            continue
    # fallback: None
    return None


def format_session_date(value):
    """Return MON-YYYY (e.g. JAN-2024) or empty string."""
    if not value:
        return ""
    # If it's a date/datetime, convert to MON-YYYY
    if hasattr(value, "strftime"):
        return value.strftime("%b-%Y").upper()
    # If it's already a string like JAN-2024 or Jan-2024 or 2024-01, normalize
    s = str(value).strip()
    # If looks like e.g. '2024-01' convert to 'JAN-2024'
    m = re.match(r"^(\d{4})[-/](\d{1,2})$", s)
    if m:
        year = m.group(1)
        month_no = int(m.group(2))
        # reverse lookup month abbrev
        month_abbr = (
            list(MONTH_MAP.keys())[month_no - 1] if 1 <= month_no <= 12 else None
        )
        if month_abbr:
            return f"{month_abbr}-{year}"
    # If already MON-YYYY (or similar), upper and normalize hyphen
    s2 = s.replace(" ", "-").replace("/", "-").upper()
    return s2


def quarter_from_date(date_str):
    if not date_str:
        return None, None
    try:
        month_str, year_str = date_str.upper().split("-")[:2]
        month = MONTH_MAP.get(month_str)
        return (f"Q{(month - 1) // 3 + 1}" if month else None), year_str
    except Exception:
        return None, None


def apply_filters(params, base_qs=None):
    """
    CRITICAL FIX: Quarterly filtering now checks trained_date and certified_date.
    If filtering by Q1 and student was trained in Q1 but certified in Q2,
    the student appears as TRAINED (not certified) when Q1 is selected.
    When Q2 is selected, the student appears as CERTIFIED (not trained).
    """
    print(params)
    qs = base_qs if base_qs is not None else studentdata.objects.all()
    center = params.get("center")
    if center:
        qs = qs.filter(center_name=center)

    mode = params.get("mode")
    if mode:
        qs = qs.filter(mode=mode)

    caste = params.get("caste")
    if caste:
        qs = qs.filter(caste_category=caste)

    trained = params.get("trained")
    if trained:
        if trained == "or":
            qs = qs.filter(Q(trained=True) | Q(certified=True))
        else:
            qs = qs.filter(trained=parse_bool(trained))

    certified = params.get("certified")
    if certified:
        qs = qs.filter(certified=parse_bool(certified))

    placed = params.get("placed")
    if placed:
        qs = qs.filter(placed=parse_bool(placed))

    claimed = params.get("claimed")
    if claimed:
        qs = qs.filter(claimed=parse_bool(claimed))

    scheme = params.get("scheme")
    if scheme:
        qs = qs.filter(scheme=scheme)

    nsqf = params.get("nsqf")
    if nsqf == "no":
        qs = qs.filter(nsqf=None)
    elif nsqf == "yes":
        qs = qs.filter(nsqf__regex=r".+")

    qmap = {
        "Q1": ["APR", "MAY", "JUN"],
        "Q2": ["JUL", "AUG", "SEP"],
        "Q3": ["OCT", "NOV", "DEC"],
        "Q4": ["JAN", "FEB", "MAR"],
    }
    quarter = params.get("quarterly")

    if quarter:
        months = qmap.get(quarter, [])
        w = Q()
        for m in months:
            # Filter by trained_date OR certified_date (not session)
            # This ensures quarterly filters check when training/certification actually occurred
            w = (
                w
                | Q(trained_date__startswith=f"{m}-")
                | Q(certified_date__startswith=f"{m}-")
            )
        qs = qs.filter(w)

    yearly = params.get("year")
    if yearly:
        qs = qs.filter(
            Q(trained_date__contains=yearly) | Q(certified_date__contains=yearly)
        )

    return qs


def student_to_dict(s, selected_quarter=None):
    # Calculate claimable amount based on selected quarter if provided
    if selected_quarter:
        claimable = float(s.get_claimable_amount_for_quarter(selected_quarter))
    else:
        claimable = float(s.claimable_amount)
    
    return {
        "dob": s.dob.isoformat() if s.dob else "",
        "fee_date": s.fee_date.isoformat() if s.fee_date else "",
        "trained_date": s.trained_date or "",
        "session": s.session or "",
        "id": s.id | 0,
        "roll_number": s.roll_number,
        "batch_code": s.batch_code,
        "name": s.name,
        "father_name": s.father_name,
        "mother_name": s.mother_name,
        "gender": s.gender,
        "address": s.address,
        "qualifications": s.qualifications,
        "aadhaar": s.aadhaar,
        "course_name": s.course_name,
        "scheme": s.scheme,
        "nsqf": s.nsqf,
        "course_hour": s.course_hour,
        "course_category": s.course_category,
        "center_name": s.center_name,
        "mode": s.mode,
        "caste_category": s.caste_category,
        "fee": float(s.fee),
        "claimable_amount": claimable,
        "trained": s.trained,
        "certified": s.certified,
        "certified_date": s.certified_date,
        "placed": s.placed,
        "claimed": s.claimed,
    }


def xlrow_to_dict(s):
    return {
        "roll_number": s.roll_number,
        "batch_code": s.batch_code,
        "name": s.name,
        "father_name": s.father_name,
        "mother_name": s.mother_name,
        "dob": s.dob.strftime("%Y-%m-%d") if s.dob else "",
        "gender": s.gender,
        "address": s.address,
        "qualifications": s.qualifications,
        "aadhaar": s.aadhaar,
        "course_name": s.course_name,
        "scheme": s.scheme,
        "nsqf": s.nsqf,
        "course_hour": s.course_hour,
        "center_name": s.center_name,
        "mode": s.mode,
        "caste_category": s.caste_category,
        "fee": float(s.fee),
        "fee_date": s.fee_date or "",
        "trained_date": s.trained_date,
        "certified_date": s.certified_date,
        "placed": s.placed,
        "claimed": s.claimed,
        "session": s.session,
    }


def center_summary(qs):
    summary = {
        "Total": qs.count(),
        "SC": 0,
        "ST": 0,
        "OBC": 0,
        "PWD": 0,
        "GENERAL": 0,
        "B": 0,
        "C": 0,
        "D": 0,
        "E": 0,
    }
    for s in qs:
        if s.caste_category in summary:
            summary[s.caste_category] += 1
        cat = (s.course_category or "").strip().upper()[:1]
        if cat in summary:
            summary[cat] += 1
    return summary


def login_view(request):
    if request.method == "POST":
        user = authenticate(
            request,
            username=request.POST["username"],
            password=request.POST["password"],
        )
        if user:
            login(request, user)
            return redirect("dashboard")
        return render(request, "login.html", {"error": "Invalid credentials"})
    return render(request, "login.html")


def logout_view(request):
    logout(request)
    return redirect("login")


def landing(request):
    if request.user.is_authenticated:
        return redirect("dashboard")
    return render(request, "landing.html")


@login_required(login_url="/login")
@ensure_csrf_cookie
def dashboard(request):
    profile = getattr(request.user, "profile", None)
    user_center = profile.center if profile and profile.center else None
    return render(request, "dashboard.html", {"user_center": user_center})


@login_required(login_url="/login")
def upload(request):
    """
    Upload Excel and create studentdata rows.
    session is taken from the upload page dropdowns (session=month, year) and stored as MON-YYYY.
    Trained/certified booleans are inferred from explicit columns or presence of trained_date/certified_date.
    If trained/certified is True but corresponding *_date is missing, set date to the selected session.
    """
    if request.method == "POST":
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES.get("file")
            month = (request.POST.get("session") or "").strip().upper()
            year = (request.POST.get("year") or "").strip()
            form_session = f"{month}-{year}" if month and year else ""

            profile = getattr(request.user, "profile", None)
            user_center = profile.center if profile and profile.center else None

            try:
                wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                ws = wb.active
                created = 0
                updated = 0
                errors = 0
                center_skipped = 0
                headers = []
                # Build header names normalized to lower-case keys
                for cell in ws[1]:
                    value = cell.value
                    headers.append(
                        str(value).lower().strip() if value is not None else ""
                    )

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))

                    row_center = str(row_dict.get("center_name") or "").strip().lower()
                    if user_center and row_center and row_center != user_center:
                        center_skipped += 1
                        continue

                    name = row_dict.get("name")
                    course = row_dict.get("course_name")
                    if not name or not course:
                        errors += 1
                        continue

                    # ── Match existing record (like placement logic) ──
                    # Primary: aadhaar + course_name + batch_code
                    # Fallback: name + course_name + batch_code
                    aadhaar = row_dict.get("aadhaar")
                    batch = row_dict.get("batch_code")
                    student = None
                    if aadhaar and course:
                        qs = studentdata.objects.filter(
                            aadhaar=str(aadhaar).strip(),
                            course_name__iexact=str(course).strip(),
                        )
                        if batch:
                            qs = qs.filter(batch_code__iexact=str(batch).strip())
                        if user_center:
                            qs = qs.filter(center_name=user_center)
                        student = qs.first()

                    if not student and name and course:
                        qs = studentdata.objects.filter(
                            name__iexact=str(name).strip(),
                            course_name__iexact=str(course).strip(),
                        )
                        if batch:
                            qs = qs.filter(batch_code__iexact=str(batch).strip())
                        if user_center:
                            qs = qs.filter(center_name=user_center)
                        student = qs.first()

                    # Booleans: prefer explicit boolean/text column, fallback to presence of date column
                    trained_bool = parse_bool(row_dict.get("trained")) or bool(
                        row_dict.get("trained_date")
                    )
                    certified_bool = parse_bool(row_dict.get("certified")) or bool(
                        row_dict.get("certified_date")
                    )
                    placed_bool = parse_bool(row_dict.get("placed"))

                    # Parse numeric fields safely
                    try:
                        course_hour_val = int(row_dict.get("course_hour") or 0)
                    except Exception:
                        course_hour_val = 0

                    try:
                        fee_val = Decimal(str(row_dict.get("fee") or 0))
                    except Exception:
                        fee_val = Decimal("0.00")

                    # Parse dates
                    dob_val = parse_date(row_dict.get("dob"))
                    fee_date_val = parse_date(row_dict.get("fee_date"))

                    # Format trained/certified dates to MON-YYYY if present
                    trained_date_val = (
                        format_session_date(row_dict.get("trained_date"))
                        if row_dict.get("trained_date")
                        else ""
                    )
                    certified_date_val = (
                        format_session_date(row_dict.get("certified_date"))
                        if row_dict.get("certified_date")
                        else ""
                    )

                    # If boolean True but date missing, set date to dropdown session
                    if trained_bool and not trained_date_val:
                        trained_date_val = form_session or trained_date_val
                    if certified_bool and not certified_date_val:
                        certified_date_val = form_session or certified_date_val

                    def safe_str(val):
                        return str(val).strip() if val is not None else ""

                    try:
                        mutable_fields = {
                            "session": form_session,
                            "roll_number": safe_str(row_dict.get("roll_number")),
                            "batch_code": safe_str(row_dict.get("batch_code")),
                            "gender": safe_str(row_dict.get("gender")),
                            "address": safe_str(row_dict.get("address")),
                            "qualifications": safe_str(row_dict.get("qualifications")),
                            "scheme": safe_str(row_dict.get("scheme")),
                            "nsqf": safe_str(row_dict.get("nsqf")),
                            "course_hour": course_hour_val,
                            "mode": safe_str(row_dict.get("mode")),
                            "caste_category": safe_str(row_dict.get("caste_category")),
                            "center_name": user_center or safe_str(row_dict.get("center_name")),
                            "fee": fee_val,
                            "fee_date": fee_date_val,
                            "trained": trained_bool,
                            "trained_date": trained_date_val,
                            "certified": certified_bool,
                            "certified_date": certified_date_val,
                            "placed": placed_bool,
                        }

                        all_fields = {
                            **mutable_fields,
                            "name": safe_str(row_dict.get("name")),
                            "father_name": safe_str(row_dict.get("father_name")),
                            "mother_name": safe_str(row_dict.get("mother_name")),
                            "dob": dob_val,
                            "aadhaar": safe_str(row_dict.get("aadhaar")),
                            "course_name": safe_str(row_dict.get("course_name")),
                        }

                        if student:
                            for field, value in mutable_fields.items():
                                setattr(student, field, value)
                            student.save()
                            updated += 1
                        else:
                            studentdata.objects.create(**all_fields)
                            created += 1
                    except Exception as e:
                        print(f"Error on row: {e}")
                        errors += 1

                skip_msg = f" | Center mismatched skipped: {center_skipped}" if center_skipped else ""
                messages.success(
                    request,
                    f"Created: {created} | Updated: {updated} | Errors: {errors}{skip_msg}",
                )
            except Exception as e:
                print(f"❌ Failed to open Excel: {e}")
                messages.error(request, f"Error processing file: {str(e)}")
    else:
        form = ExcelUploadForm()

    return render(request, "upload.html", {"form": form})


@login_required(login_url="/login")
def filter_students(request):
    students = apply_filters(request.GET, get_student_qs(request.user))
    selected_quarter = request.GET.get("quarterly")  # Get selected quarter for claimable amount calculation
    page = int(request.GET.get("page", 1))
    limit = int(request.GET.get("limit", 10))
    offset = (page - 1) * limit
    total = len(students) if isinstance(students, list) else students.count()

    return JsonResponse(
        {
            "results": [student_to_dict(s, selected_quarter) for s in students[offset : offset + limit]],
            "pagination": {
                "page": page,
                "limit": limit,
                "total_count": total,
                "total_pages": (total + limit - 1) // limit,
                "has_next": page < (total + limit - 1) // limit,
                "has_prev": page > 1,
            },
        }
    )


@login_required(login_url="/login")
def download_filtered_data(request):
    try:
        students = apply_filters(request.GET, get_student_qs(request.user))
        selected_quarter = request.GET.get("quarterly")  # Get selected quarter for claimable amount calculation
        print(f"DEBUG: students count {len(students) if isinstance(students, list) else students.count()}, quarter: {selected_quarter}")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Filtered Students"

        headers = [
            "Roll Number",
            "Batch Code",
            "Name",
            "Father Name",
            "Mother Name",
            "DOB",
            "Gender",
            "Address",
            "Qualifications",
            "Aadhaar",
            "Course Name",
            "Scheme",
            "NSQF",
            "Course Hours",
            "Course Category",
            "Center",
            "Mode",
            "Caste Category",
            "Fee",
            "Claimable Amount",
            "Fee Date",
            "Trained",
            "Trained Date",
            "Certified",
            "Certified Date",
            "Placed",
            "Claimed",
            "Session",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        yn = lambda v: "Yes" if v else "No"
        for row, s in enumerate(students, 2):
            # Calculate claimable amount based on selected quarter if provided
            if selected_quarter:
                claimable = s.get_claimable_amount_for_quarter(selected_quarter)
            else:
                claimable = s.claimable_amount
            
            for col, val in enumerate(
                [
                    s.roll_number,
                    s.batch_code,
                    s.name,
                    s.father_name,
                    s.mother_name,
                    s.dob.strftime("%Y-%m-%d") if s.dob else "",
                    s.gender,
                    s.address,
                    s.qualifications,
                    s.aadhaar,
                    s.course_name,
                    s.scheme,
                    s.nsqf,
                    s.course_hour,
                    s.course_category,
                    s.center_name,
                    s.mode,
                    s.caste_category,
                    float(s.fee),
                    float(claimable),
                    str(s.fee_date) if s.fee_date else "",
                    yn(s.trained),
                    s.trained_date,
                    yn(s.certified),
                    s.certified_date,
                    yn(s.placed),
                    yn(s.claimed),
                    s.session,
                ],
                1,
            ):
                ws.cell(row=row, column=col, value=val)

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max((len(str(c.value)) for c in col if c.value), default=0) + 2, 50
            )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=filtered_students.xlsx"
        wb.save(response)
        return response

    except Exception as e:
        print(f"DEBUG: Error in download_filtered_data: {e}")
        print(traceback.format_exc())
        return HttpResponse("Error processing download", status=500)


# ─── Report ──────────────────────────────────────────────────────────────────


def _session_filter_options(base_qs=None):
    qs = base_qs if base_qs is not None else studentdata.objects.all()
    sessions = list(
        qs.values_list("session", flat=True)
        .distinct()
        .order_by("-session")
    )
    years = sorted({s.split("-")[1] for s in sessions if "-" in s}, reverse=True)
    months = sorted({s.split("-")[0] for s in sessions if "-" in s})
    return years, months


@login_required(login_url="/login")
def download(request):
    p = request.GET
    students = get_student_qs(request.user)
    for key in ["year", "session"]:
        if p.get(key):
            students = students.filter(session__icontains=p[key])
    if p.get("center"):
        students = students.filter(center_name=p["center"])

    grouped = {}
    for s in students:
        key = f"{s.course_category}|{s.course_name}|{s.center_name}|{s.session}"
        if key not in grouped:
            grouped[key] = {
                "category": s.course_category,
                "course_name": s.course_name,
                "course_hour": s.course_hour,
                "center_name": s.center_name,
                "session": s.session,
                "scheme": s.scheme,
                "nsqf": s.nsqf,
                **{
                    c: {"trained": 0, "certified": 0, "placed": 0, "total": 0}
                    for c in ["GENERAL", "OBC", "SC", "ST", "PWD"]
                },
            }
        caste = s.caste_category if s.caste_category in grouped[key] else "GENERAL"
        g = grouped[key][caste]
        g["total"] += 1
        if s.trained_date:
            g["trained"] += 1
        if s.certified_date:
            g["certified"] += 1
        if s.placed:
            g["placed"] += 1

    report_data = list(grouped.values())
    castes = ["GENERAL", "OBC", "SC", "ST", "PWD"]
    totals = {
        c: {"trained": 0, "certified": 0, "placed": 0, "total": 0} for c in castes
    }
    for item in report_data:
        for c in castes:
            for k in totals[c]:
                totals[c][k] += item[c][k]
    totals["grand_total"] = sum(totals[c]["total"] for c in castes)

    years, months = _session_filter_options(get_student_qs(request.user))
    return render(
        request,
        "download.html",
        {
            "data": report_data,
            "totals": totals,
            "selected_year": p.get("year", ""),
            "selected_session": p.get("session", ""),
            "selected_center": p.get("center", ""),
            "years": years,
            "months": months,
            "centers": CENTERS,
        },
    )


@login_required(login_url="/login")
def api_download_data(request):
    p = request.GET
    students = get_student_qs(request.user)
    for key in ["year", "session"]:
        if p.get(key):
            students = students.filter(session__icontains=p[key])
    if p.get("center"):
        students = students.filter(center_name=p["center"])

    return JsonResponse(
        {
            "results": [
                {
                    "course_category": s.course_category,
                    "course_name": s.course_name,
                    "course_hour": s.course_hour,
                    "center_name": s.center_name,
                    "scheme": s.scheme,
                    "nsqf": s.nsqf,
                    "session": s.session,
                    "caste_category": s.caste_category,
                    "trained_date": s.trained_date,
                    "certified_date": s.certified_date,
                    "placed": s.placed,
                    "fee": float(s.fee),
                    "claimable_amount": float(s.claimable_amount),
                }
                for s in students
            ]
        }
    )


# ─── Update ──────────────────────────────────────────────────────────────────


@login_required(login_url="/login")
def update_student(request, student_id):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    try:
        student = studentdata.objects.get(id=student_id)
    except studentdata.DoesNotExist:
        return JsonResponse({"error": "Student not found"}, status=404)

    profile = getattr(request.user, "profile", None)
    if profile and profile.center and student.center_name != profile.center:
        return JsonResponse({"error": "Access denied"}, status=403)

    try:
        body = json.loads(request.body)
        current_month = (
            datetime.now().strftime("%b").upper() + "-" + datetime.now().strftime("%Y")
        )

        str_fields = [
            "name",
            "father_name",
            "mother_name",
            "address",
            "qualifications",
            "aadhaar",
            "course_name",
            "scheme",
            "nsqf",
        ]
        for f in str_fields:
            if body.get(f) is not None:
                setattr(student, f, str(body[f]).strip())

        session_value = body.get("session")
        if session_value is not None:
            session_value = str(session_value).strip().upper()
            if re.fullmatch(
                r"(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)-\d{4}",
                session_value,
            ):
                student.session = session_value

        if body.get("batch_code"):
            student.batch_code = str(body["batch_code"]).strip().upper()

        for f in ["gender", "mode", "caste_category", "center_name"]:
            if body.get(f) is not None:
                setattr(student, f, body[f])

        for f in ["dob", "fee_date"]:
            val = body.get(f)
            if val is not None:
                setattr(student, f, val if val else None)

        try:
            student.course_hour = int(
                body.get("course_hour") or student.course_hour or 0
            )
        except (ValueError, TypeError):
            student.course_hour = 0

        try:
            student.fee = float(body.get("fee") or student.fee or 0)
        except (ValueError, TypeError):
            student.fee = 0.0

        if body.get("placed") is not None:
            student.placed = body["placed"]

        # Inside update_student view:
        current_session_label = datetime.now().strftime("%b-%Y").upper()

        for field in ["trained", "certified"]:
            new_val = body.get(field)
            if new_val is not None:
                date_field = f"{field}_date"
                date_val = body.get(date_field)
                if new_val:
                    if date_val:
                        setattr(student, date_field, format_session_date(date_val))
                    elif not getattr(student, date_field):
                        setattr(student, date_field, current_session_label)
                else:
                    setattr(student, date_field, "")
                setattr(student, field, new_val)

        if body.get("claimed") is not None:
            student.claimed = body["claimed"]

        student.save()
        return JsonResponse(
            {
                "success": True,
                "course_category": student.course_category,
                "claimable_amount": float(student.claimable_amount),
                "trained_date": student.trained_date,
                "certified_date": student.certified_date,
                "claimed": student.claimed,
            }
        )

    except json.JSONDecodeError as e:
        return JsonResponse(
            {"success": False, "error": f"Invalid JSON: {e}"}, status=400
        )
    except Exception as e:
        print(traceback.format_exc())
        return JsonResponse({"success": False, "error": str(e)}, status=400)


@login_required(login_url="/login")
def inputView(request):
    if request.method == "POST":
        form = StudentDataForm(request.POST)
        if form.is_valid():
            student = form.save(commit=False)
            if student.trained and not student.trained_date:
                student.trained_date = student.session
            if student.certified and not student.certified_date:
                student.certified_date = student.session
            student.save()
            return redirect("dashboard")
    else:
        form = StudentDataForm()

    return render(
        request,
        "input.html",
        {
            "form": form,
            "months": [
                "JAN",
                "FEB",
                "MAR",
                "APR",
                "MAY",
                "JUN",
                "JUL",
                "AUG",
                "SEP",
                "OCT",
                "NOV",
                "DEC",
            ],
            "years": list(range(2020, 2031)),
        },
    )


# ─── Overview ────────────────────────────────────────────────────────────────


def _overview_context(selected_session, user=None):
    base_qs = get_student_qs(user) if user else studentdata.objects.all()
    students = base_qs
    if selected_session:
        students = students.filter(session=selected_session)

    return {
        "all_record": students.count(),
        "centers": [
            {
                "name": n.capitalize(),
                "stats": center_summary(students.filter(center_name=n)),
            }
            for n in CENTERS
        ],
        "sessions": list(
            base_qs.values_list("session", flat=True)
            .distinct()
            .order_by("-session")
        ),
        "selected_session": selected_session,
    }


@login_required(login_url="/login")
def overview(request):
    ctx = _overview_context(request.GET.get("session", ""), user=request.user)
    ctx["centers"] = [(c["name"], c["stats"]) for c in ctx["centers"]]
    return render(request, "overview.html", ctx)


@login_required(login_url="/login")
def overview_data(request):
    return JsonResponse(_overview_context(request.GET.get("session", ""), user=request.user))


def courses(request):
    return render(
        request,
        "view_courses.html",
        {
            "It": NsqfIT.objects.all(),
            "elctro": NsqfElectronics.objects.all(),
            "dlc": Dlc.objects.all(),
        },
    )


def add_dropdown(field_name, option_list):
    col = column_index[field_name]
    col_letter = get_column_letter(col)
    # Build the formula: e.g., '"offline,online"'
    formula = '"' + ",".join(option_list) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Please pick a value from the list."
    dv.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv)
    # Apply to all rows from 2 to 1,048,576
    dv.add(col_letter + "2:" + col_letter + "1048576")


import openpyxl
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import studentdata


@login_required(login_url="/login")
def sample_upload(request):

    headers = [
        "roll_number",
        "batch_code",
        "name",
        "father_name",
        "mother_name",
        "dob",
        "gender",
        "address",
        "qualifications",
        "aadhaar",
        "course_name",
        "scheme",
        "nsqf",
        "course_hour",
        "mode",
        "caste_category",
        "center_name",
        "fee",
        "fee_date",
        "trained_date",
        "certified_date",
        "placed",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Data Template"

    col_num = 1
    for header in headers:
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        col_num = col_num + 1

    ws.freeze_panes = "A2"

    column_index = {}
    idx = 1
    for h in headers:
        column_index[h] = idx
        idx = idx + 1

    def add_dropdown(field_name, options_list):
        col_num = column_index[field_name]
        col_letter = get_column_letter(col_num)

        formula = '"' + ",".join(options_list) + '"'

        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Please select a value from the dropdown list."
        dv.errorTitle = "Invalid Entry"
        ws.add_data_validation(dv)

        dv.add(f"{col_letter}2:{col_letter}1048576")

    schemes_qs = get_student_qs(request.user).values_list("scheme", flat=True).distinct()

    scheme_options = []
    for s in schemes_qs:
        if s != None:
            scheme_options.append(s)
        else:
            scheme_options.append("NON-NSQF")

    centers = []
    for i in studentdata.CENTER_CHOICES:
        centers.append(i[1])

    caste_choices = []
    for i in studentdata.CASTE_CHOICES:
        caste_choices.append(i[1])

    mode_choices = []
    for i in studentdata.MODE_CHOICES:
        mode_choices.append(i[1])

    gender_choices = []
    for i in studentdata.GENDER:
        gender_choices.append(i[1])

    nsqf_choices = []
    for i in studentdata.NSQF_LEVEL:
        nsqf_choices.append(i[1])

    course_choioces = []

    for i in studentdata.COURSE_CHOICES:
        course_choioces.append(i[1])

    qualification_choices = []
    for i in studentdata.HIGHEST_QUALIFICATION:
        qualification_choices.append(i[1])
    add_dropdown("center_name", centers)
    add_dropdown("mode", mode_choices)
    add_dropdown("caste_category", caste_choices)
    add_dropdown("placed", ["True", "False"])
    add_dropdown("gender", gender_choices)
    add_dropdown("nsqf", nsqf_choices)
    add_dropdown("scheme", scheme_options)
    add_dropdown("course_name", course_choioces)
    add_dropdown("qualifications", qualification_choices)

    sample_row = [
        "NIELIT0001",
        "Batch-2024-001",
        "Aarav Sharma",
        "Rajesh Sharma",
        "Sunita Sharma",
        "1998-05-12",
        "Male",
        "12/4 Rohini Delhi",
        "Graduation",
        "123456789012",
        "Python Programming",
        "PMKVY",
        "Level 4",
        "120",
        "offline",
        "OBC",
        "inderlok",
        "5000",
        "2024-01-10",
        "JAN-2024",
        "MAR-2024",
        "TRUE",
    ]
    sample_col = 1
    for value in sample_row:
        ws.cell(row=2, column=sample_col, value=value)
        sample_col = sample_col + 1

    col = 1
    while col <= len(headers):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18
        col = col + 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="student_upload_template.xlsx"'
    )
    wb.save(response)
    return response


PLACEMENT_COLUMN_ALIASES = {
    "student name": "name",
    "student id": "aadhaar",
    "course/branch": "course_name",
    "semester/batch": "batch_code",
    "centre/campus": "center_name",
    "offer received (y/n)": "offer_received",
    "selection status": "selection_status",
    "opportunity type (internship/placement)": "opportunity_type",
    "company/organization": "company",
    "job title/role": "job_title",
    "source of opportunity": "source",
    "date applied": "date_applied",
    "joining date": "joining_date",
    "current status": "current_status",
    "s. no.": "s_no",
}


def _normalize_placement_headers(headers):
    norm = []
    for h in headers:
        key = h.lower().strip().replace("\n", " ").replace("\r", " ")
        key = re.sub(r"\s+", " ", key)
        norm.append(PLACEMENT_COLUMN_ALIASES.get(key, key))
    return norm


def _parse_offer_received(val):
    if val is None:
        return None
    s = str(val).strip().upper()
    return s in ("Y", "YES", "TRUE", "1")


def _parse_selection_status(val):
    if val is None:
        return None
    s = str(val).strip().lower()
    return s in ("selected", "offered", "placed")


@login_required(login_url="/login")
def upload_placement_records(request):
    profile = getattr(request.user, "profile", None)
    user_center = profile.center if profile and profile.center else None

    if request.method == "POST":
        form = PlacementUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES.get("file")
            try:
                wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                ws = wb.active
                created = 0
                updated = 0
                not_found = 0
                errors = 0
                skipped = 0

                raw_headers = []
                for cell in ws[1]:
                    value = cell.value
                    raw_headers.append(
                        str(value).strip() if value is not None else ""
                    )
                headers = _normalize_placement_headers(raw_headers)

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))

                    placed_bool = None
                    offer = _parse_offer_received(row_dict.get("offer_received"))
                    if offer is not None:
                        placed_bool = offer
                    else:
                        status = _parse_selection_status(row_dict.get("selection_status"))
                        if status is not None:
                            placed_bool = status

                    if placed_bool is None:
                        skipped += 1
                        continue

                    def csv(val):
                        return str(val).strip() if val is not None else ""

                    try:
                        student = None
                        aadhaar_val = row_dict.get("aadhaar")
                        if aadhaar_val:
                            qs = studentdata.objects.filter(aadhaar=str(aadhaar_val).strip())
                            if user_center:
                                qs = qs.filter(center_name=user_center)
                            student = qs.first()

                        if not student:
                            name_val = row_dict.get("name")
                            batch_val = row_dict.get("batch_code")
                            center_val = row_dict.get("center_name")
                            if name_val and batch_val:
                                qs = studentdata.objects.filter(
                                    name__iexact=str(name_val).strip(),
                                    batch_code__iexact=str(batch_val).strip(),
                                )
                                if user_center:
                                    qs = qs.filter(center_name=user_center)
                                elif center_val:
                                    qs = qs.filter(center_name=str(center_val).strip().lower())
                                student = qs.first()

                        # Update placed on studentdata if found
                        if student:
                            student.placed = placed_bool
                            student.save()

                        # ── Match or create PlacementRecord ──
                        match_qs = PlacementRecord.objects.filter(
                            student_name__iexact=csv(row_dict.get("name")),
                            aadhaar=csv(row_dict.get("aadhaar")),
                            course_name__iexact=csv(row_dict.get("course_name")),
                        )
                        if user_center:
                            match_qs = match_qs.filter(center_name=user_center)

                        pr = match_qs.first()
                        pr_data = {
                            "student": student,
                            "student_name": csv(row_dict.get("name")),
                            "aadhaar": csv(row_dict.get("aadhaar")),
                            "course_name": csv(row_dict.get("course_name")),
                            "batch_code": csv(row_dict.get("batch_code")),
                            "center_name": user_center or csv(row_dict.get("center_name")),
                            "opportunity_type": csv(row_dict.get("opportunity_type")),
                            "selection_status": csv(row_dict.get("selection_status")),
                            "offer_received": placed_bool,
                            "company": csv(row_dict.get("company")),
                            "job_title": csv(row_dict.get("job_title")),
                            "source": csv(row_dict.get("source")),
                            "date_applied": csv(row_dict.get("date_applied")),
                            "joining_date": csv(row_dict.get("joining_date")),
                            "current_status": csv(row_dict.get("current_status")),
                            "placed": placed_bool,
                        }

                        if pr:
                            for field, value in pr_data.items():
                                setattr(pr, field, value)
                            pr.save()
                            updated += 1
                        else:
                            PlacementRecord.objects.create(**pr_data)
                            created += 1
                    except Exception:
                        errors += 1

                parts = [f"Created: {created} | Updated: {updated}"]
                if not_found:
                    parts.append(f"Student link not found: {not_found}")
                if skipped:
                    parts.append(f"Skipped (no placement status): {skipped}")
                if errors:
                    parts.append(f"Errors: {errors}")
                messages.success(request, " | ".join(parts))
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
        else:
            messages.error(request, "Invalid form submission.")
    else:
        form = PlacementUploadForm()

    return placement_view(request, upload_form=form)


@login_required(login_url="/login")
def placement_view(request, upload_form=None):
    profile = getattr(request.user, "profile", None)
    user_center = profile.center if profile and profile.center else None
    if upload_form is None:
        upload_form = PlacementUploadForm()
    return render(request, "placement.html", {"user_center": user_center, "upload_form": upload_form})


def placement_record_to_dict(r):
    return {
        "id": r.id,
        "student_name": r.student_name or "-",
        "aadhaar": r.aadhaar or "-",
        "course_name": r.course_name or "-",
        "batch_code": (r.batch_code or "").upper() or "-",
        "center_name": r.center_name or "-",
        "opportunity_type": r.opportunity_type or "-",
        "selection_status": r.selection_status or "-",
        "company": r.company or "-",
        "job_title": r.job_title or "-",
        "source": r.source or "-",
        "date_applied": r.date_applied or "-",
        "joining_date": r.joining_date or "-",
        "current_status": r.current_status or "-",
        "placed": r.placed,
        "created_at": r.created_at.isoformat() if r.created_at else "",
    }


@login_required(login_url="/login")
def filter_placement(request):
    profile = getattr(request.user, "profile", None)
    user_center = profile.center if profile and profile.center else None
    base_qs = PlacementRecord.objects.all()
    if user_center:
        base_qs = base_qs.filter(center_name=user_center)

    batch = request.GET.get("batch", "").strip()
    name = request.GET.get("name", "").strip()
    course = request.GET.get("course", "").strip()
    center = request.GET.get("center", "").strip()
    placed = request.GET.get("placed", "").strip()
    company = request.GET.get("company", "").strip()
    opportunity = request.GET.get("opportunity", "").strip()

    if batch:
        base_qs = base_qs.filter(batch_code__icontains=batch)
    if name:
        base_qs = base_qs.filter(student_name__icontains=name)
    if course:
        base_qs = base_qs.filter(course_name__icontains=course)
    if center:
        base_qs = base_qs.filter(center_name=center)
    if placed:
        base_qs = base_qs.filter(placed=parse_bool(placed))
    if company:
        base_qs = base_qs.filter(company__icontains=company)
    if opportunity:
        base_qs = base_qs.filter(opportunity_type__iexact=opportunity)

    page = int(request.GET.get("page", 1))
    limit = int(request.GET.get("limit", 10))
    offset = (page - 1) * limit
    total = base_qs.count()

    results = [placement_record_to_dict(r) for r in base_qs[offset : offset + limit]]

    return JsonResponse({
        "results": results,
        "pagination": {
            "page": page,
            "limit": limit,
            "total_count": total,
            "total_pages": (total + limit - 1) // limit if limit > 0 else 0,
            "has_next": page < (total + limit - 1) // limit if limit > 0 else False,
            "has_prev": page > 1,
        },
    })
